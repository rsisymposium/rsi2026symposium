"""
Build RSI_2026_Schedule.xlsx, the master editable spreadsheet.

Two sheets:
  - "Data": one row per student (name, title, mentor, field, pronunciation,
    email). This is a lookup table, not meant to be reformatted.
  - "Schedule": a timetable grid -- one row per time slot, one column-group
    per room (concurrent rooms side by side), one section per block (T1-T4
    Thursday, F1-F2 Friday -- see lib_scheduling.py for why the day is
    shaped this way). You type/change a student's name in a room's Name
    cell; the Field/Title/Mentor cells next to it are formulas (VLOOKUP
    into Data) and update automatically.

Initial Name cells are pre-filled using the mentor-preference-aware
placement in lib_scheduling.py (MENTOR_BLOCK_OVERRIDES etc.) -- a draft
reflecting the mentor scheduling-preferences form, not a final answer. Feel
free to drag names around; the formulas will keep up.

After hand-editing the Schedule sheet, run 03_export_site_csv.py to refresh
docs/data/schedule.csv, which is what docs/index.html actually reads.

Rerun this script only if you want to throw away hand-edits and regenerate
a fresh placement draft (e.g. the roster or preference data changed a lot).
"""
import csv
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from lib_scheduling import (
    BLOCK_BY_CODE, BLOCKS, BREAK_AFTER, DAYS, FIELD_COLORS_HEX, ROOMS, pack,
    slot_times,
)

ROOT = Path(__file__).resolve().parent.parent
ROSTER_PATH = ROOT / "data" / "roster.csv"
OUT_PATH = ROOT / "RSI_2026_Schedule.xlsx"

DATA_HEADERS = ["Full Name", "Title", "Mentor", "Field", "First Pron", "Last Pron", "Email", "Kerb"]

ROOM_COL_START = {room: 2 + 4 * i for i, room in enumerate(ROOMS)}  # 32-124->B, 32-141->F, 32-155->J
SUBHEADERS = ["Name", "Field", "Title", "Mentor"]

HEADER_FILL = PatternFill("solid", fgColor="4F46E5")
DAY_FILL = PatternFill("solid", fgColor="1E293B")
BLOCK_FILL = PatternFill("solid", fgColor="E2E8F0")
BREAK_FILL = PatternFill("solid", fgColor="FEF3C7")
SUBHEADER_FILL = PatternFill("solid", fgColor="F1F5F9")
THIN = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_roster():
    with open(ROSTER_PATH, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def build_data_sheet(wb, students):
    ws = wb.active
    ws.title = "Data"
    ws.append(DATA_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for s in students:
        ws.append([
            s["full_name"], s["title"], s["mentor"], s["field"],
            s["first_pron"], s["last_pron"], s["email"], s["kerb"],
        ])
    widths = [24, 46, 26, 26, 20, 20, 22, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(students) + 1}"
    return len(students)


def merge_row(ws, row, col_start, col_end, value, font=None, fill=None, align=None):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = align or Alignment(horizontal="center", vertical="center")
    return cell


def build_schedule_sheet(wb, state):
    ws = wb.create_sheet("Schedule")
    total_cols = 1 + 4 * len(ROOMS)

    merge_row(ws, 1, 1, total_cols, "RSI 2026 Symposium Schedule",
              font=Font(bold=True, size=16, color="1E293B"))
    merge_row(ws, 2, 1, total_cols,
              "Edit a Name cell to reassign a slot -- Field / Title / Mentor update automatically.",
              font=Font(italic=True, size=10, color="64748B"))

    row = 4
    name_ranges = []

    for day in DAYS:
        row += 1
        merge_row(ws, row, 1, total_cols, day["display"],
                  font=Font(bold=True, size=13, color="FFFFFF"), fill=DAY_FILL)
        row += 1

        for code in day["block_codes"]:
            block = BLOCK_BY_CODE[code]
            start_label = datetime.strptime(block["start"], "%H:%M").strftime("%-I:%M %p")
            end_label = datetime.strptime(block["end"], "%H:%M").strftime("%-I:%M %p")
            label = f"Block {code} ({start_label} – {end_label})"
            merge_row(ws, row, 1, total_cols, label,
                      font=Font(bold=True, size=11, color="1E293B"), fill=BLOCK_FILL)
            row += 1

            ws.cell(row=row, column=1, value="Time").font = Font(bold=True)
            ws.cell(row=row, column=1).fill = SUBHEADER_FILL
            for room in ROOMS:
                start_col = ROOM_COL_START[room]
                merge_row(ws, row, start_col, start_col + 3, f"Room {room}",
                          font=Font(bold=True, color="FFFFFF"), fill=HEADER_FILL)
            row += 1

            for room in ROOMS:
                start_col = ROOM_COL_START[room]
                for j, sub_label in enumerate(SUBHEADERS):
                    c = ws.cell(row=row, column=start_col + j, value=sub_label)
                    c.font = Font(bold=True, size=9)
                    c.fill = SUBHEADER_FILL
                    c.alignment = Alignment(horizontal="center")
            row += 1

            capacity = block["cap"]
            times = slot_times(block["start"], capacity)
            first_data_row = row

            for i in range(capacity):
                start_label, end_label = times[i]
                ws.cell(row=row, column=1, value=f"{start_label} – {end_label}").font = Font(size=9)
                for room in ROOMS:
                    start_col = ROOM_COL_START[room]
                    name_col_letter = get_column_letter(start_col)
                    roster_here = state[code][room]
                    prefill = roster_here[i]["full_name"] if i < len(roster_here) else ""
                    name_cell = ws.cell(row=row, column=start_col, value=prefill)
                    name_cell.border = BORDER
                    field_cell = ws.cell(
                        row=row, column=start_col + 1,
                        value=f'=IFERROR(VLOOKUP(${name_col_letter}{row},Data!$A:$D,4,FALSE),"")')
                    title_cell = ws.cell(
                        row=row, column=start_col + 2,
                        value=f'=IFERROR(VLOOKUP(${name_col_letter}{row},Data!$A:$D,2,FALSE),"")')
                    mentor_cell = ws.cell(
                        row=row, column=start_col + 3,
                        value=f'=IFERROR(VLOOKUP(${name_col_letter}{row},Data!$A:$D,3,FALSE),"")')
                    for c in (field_cell, title_cell, mentor_cell):
                        c.border = BORDER
                        c.alignment = Alignment(wrap_text=True, vertical="center")
                        c.font = Font(size=9)
                row += 1

            last_data_row = row - 1
            for room in ROOMS:
                name_col_letter = get_column_letter(ROOM_COL_START[room])
                name_ranges.append((name_col_letter, first_data_row, last_data_row))

            if code in BREAK_AFTER:
                merge_row(ws, row, 1, total_cols, BREAK_AFTER[code],
                          font=Font(bold=True, italic=True, color="92400E"), fill=BREAK_FILL)
                row += 1
        row += 1  # gap between days

    ws.column_dimensions["A"].width = 20
    for room in ROOMS:
        start_col = ROOM_COL_START[room]
        ws.column_dimensions[get_column_letter(start_col)].width = 22
        ws.column_dimensions[get_column_letter(start_col + 1)].width = 15
        ws.column_dimensions[get_column_letter(start_col + 2)].width = 42
        ws.column_dimensions[get_column_letter(start_col + 3)].width = 22

    num_students = ws.parent["Data"].max_row - 1
    dv = DataValidation(type="list", formula1=f"=Data!$A$2:$A${num_students + 1}",
                         allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    for col_letter, first_row, last_row in name_ranges:
        dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")

    field_cols = [get_column_letter(ROOM_COL_START[room] + 1) for room in ROOMS]
    max_row = row + 5
    for field_name, (bg, fg) in FIELD_COLORS_HEX.items():
        fill = PatternFill("solid", fgColor=bg)
        font = Font(color=fg, size=9)
        for col_letter in field_cols:
            ws.conditional_formatting.add(
                f"{col_letter}4:{col_letter}{max_row}",
                CellIsRule(operator="equal", formula=[f'"{field_name}"'], fill=fill, font=font),
            )

    ws.sheet_view.showGridLines = False
    return ws


def main():
    students = load_roster()
    state, warnings = pack(students)

    wb = Workbook()
    n = build_data_sheet(wb, students)
    build_schedule_sheet(wb, state)

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH} ({n} students)")

    placed = sum(len(state[b["code"]][room]) for b in BLOCKS for room in ROOMS)
    print(f"Placed {placed}/{n} students.")
    print("\nBlock sizes:")
    for b in BLOCKS:
        code = b["code"]
        for room in ROOMS:
            count = len(state[code][room])
            fields = sorted(set(s["field"] for s in state[code][room]))
            print(f"  {code} {room}: {count}/{b['cap']} -- {', '.join(fields)}")
    if warnings:
        print(f"\n{len(warnings)} WARNING(S):")
        for w in warnings:
            print(f"  {w}")
    if placed < n:
        unplaced = n - placed
        print(f"\nWARNING: {unplaced} student(s) were not placed anywhere -- check capacity.")


if __name__ == "__main__":
    main()
