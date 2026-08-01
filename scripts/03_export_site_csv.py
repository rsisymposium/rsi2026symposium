"""
Read RSI_2026_Schedule.xlsx (after any hand-edits) and flatten its Schedule
sheet into docs/data/schedule.csv, the long-format file docs/index.html
actually fetches and renders.

docs/ is the only folder GitHub Pages publishes -- it intentionally
contains no email addresses or other PII. Keep it that way: don't add the
raw source spreadsheets, RSI_2026_Schedule.xlsx, or data/roster.csv (all of
which contain student emails) into docs/.

This does NOT trust Excel's cached formula results (openpyxl can't evaluate
formulas). It re-does the Name -> Title/Mentor/Field lookup itself, straight
from the Data sheet.

Row/column positions are discovered by scanning for marker text ("Block
T1 (...)" headers and "H:MM AM – H:MM PM" time labels) rather than assumed
from fixed offsets, so this keeps working even if rows get inserted/deleted
while hand-editing the Schedule sheet.

Run this any time after editing RSI_2026_Schedule.xlsx to republish the
site's data.
"""
import csv
import re
from pathlib import Path

from openpyxl import load_workbook

from lib_scheduling import BLOCK_BY_CODE, DAYS, ROOMS

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "RSI_2026_Schedule.xlsx"
OUT_PATH = ROOT / "docs" / "data" / "schedule.csv"

ROOM_COL_START = {room: 2 + 4 * i for i, room in enumerate(ROOMS)}  # must match 02_build_schedule_workbook.py
DATE_TO_DAY = {d["date"]: d for d in DAYS}
BLOCK_CODE_FOR_DATE = {d["date"]: set(d["block_codes"]) for d in DAYS}
BLOCK_LABEL_RE = re.compile(r"^Block (\w+) \(")


def load_data_lookup(wb):
    ws = wb["Data"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    lookup = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[idx["Full Name"]]
        if not name:
            continue
        lookup[str(name).strip()] = {
            "title": row[idx["Title"]] or "",
            "mentor": row[idx["Mentor"]] or "",
            "field": row[idx["Field"]] or "",
        }
    return lookup


def looks_like_time_range(value):
    return isinstance(value, str) and " – " in value and ("AM" in value or "PM" in value)


def find_day_for_block_code(code):
    for d in DAYS:
        if code in d["block_codes"]:
            return d
    return None


def main():
    wb = load_workbook(XLSX_PATH, data_only=False)
    data_lookup = load_data_lookup(wb)
    ws = wb["Schedule"]

    grouped = {}  # (date, block_code, room) -> list of row dicts, in sheet order
    current_block_code = None
    current_day = None
    unknown_names = []

    for row in range(1, ws.max_row + 1):
        a_val = ws.cell(row=row, column=1).value
        if not isinstance(a_val, str):
            continue

        m = BLOCK_LABEL_RE.match(a_val)
        if m:
            current_block_code = m.group(1)
            current_day = find_day_for_block_code(current_block_code)
            continue

        if looks_like_time_range(a_val) and current_block_code and current_day:
            start_time, end_time = [p.strip() for p in a_val.split(" – ", 1)]
            for room in ROOMS:
                name_val = ws.cell(row=row, column=ROOM_COL_START[room]).value
                if not name_val or not str(name_val).strip():
                    continue
                name = str(name_val).strip()
                info = data_lookup.get(name)
                if not info:
                    unknown_names.append((name, current_day["display"], current_block_code, room))
                    continue
                key = (current_day["date"], current_block_code, room)
                grouped.setdefault(key, []).append({
                    "date": current_day["date"],
                    "weekday": current_day["weekday"],
                    "block": current_block_code,
                    "room": room,
                    "start_time": start_time,
                    "end_time": end_time,
                    "full_name": name,
                    "title": info["title"],
                    "mentor": info["mentor"],
                    "field": info["field"],
                })

    all_rows = []
    for key, rows in grouped.items():
        for i, r in enumerate(rows):
            r["order"] = i + 1
            all_rows.append(r)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "date", "weekday", "block", "room", "start_time", "end_time",
        "order", "full_name", "title", "mentor", "field",
    ]
    with open(OUT_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in all_rows:
            writer.writerow({k: r[k] for k in fieldnames})

    print(f"Wrote {len(all_rows)} talks across {len(grouped)} room-blocks to {OUT_PATH}")
    if unknown_names:
        print(f"\nWARNING: {len(unknown_names)} name(s) in the Schedule sheet were not found in the Data sheet "
              f"(typo, or not yet added?) -- these were skipped:")
        for name, day, block, room in unknown_names:
            print(f"  '{name}' -- {day} {block} {room}")


if __name__ == "__main__":
    main()
